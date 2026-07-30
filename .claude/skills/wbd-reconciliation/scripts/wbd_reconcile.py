"""
WBD Digital Calendar vs Release Schedule Reconciliation
========================================================
Usage:
    python3 wbd_reconcile.py <digital_calendar.xlsx> <release_schedule.xlsx|.csv>

The Release Schedule is the "New Releases" tab of the WBD RS Google Sheet.
Step 1 of the skill fetches it fresh from Google Drive on every run.

Preferred: pass the RS as an .xlsx (the script reads the "New Releases" sheet
directly). A .csv is also accepted for backward compatibility.

-----------------------------------------------------------------------------
IMPORTANT — current RS "New Releases" layout (0-indexed):
    0: MVPD check (go_live flag)
    1: Vendor ID
    2: Upcoming Releases (title)
    3: Type (Film / TV / Film Bundle / TV Boxset ...)
    4: Release type (Premium | Premium Reprice | Standard | Pre-Order | 4K Release)
    5: Launch Date            <-- SINGLE date per release-type row
    6: Launch Covered
    7: Avail on all platforms?
    8: Added to Title List
    9: Day of Week

The sheet used to carry separate EST (col 7) and VOD (col 9) date columns.
It no longer does: there is now ONE "Launch Date" per row. In the current
WBD model EST and VOD launch on the same day for a given release type, so the
single Launch Date is used as BOTH the EST and VOD reference when matching
against the DC.

DC column order (0-indexed):
    0: title
    2: PEST  (MM/DD/YYYY text)
    4: PVOD  (MM/DD/YYYY text)
    6: EST   (MM/DD/YYYY text)
    8: VOD   (MM/DD/YYYY text)

Rules:
    - PEST/PVOD in DC must match each other
    - PEST/PVOD in DC -> RS Premium Launch Date
    - EST/VOD  in DC -> RS Standard Launch Date
    - Only "Premium" rows (not "Premium Reprice") match PEST/PVOD
    - Only "Standard" rows match EST/VOD
    - Flag DC entries missing from RS entirely
-----------------------------------------------------------------------------
"""

import sys
import csv
import openpyxl
from datetime import datetime, date
import re
from collections import defaultdict

RS_SHEET_NAME = "New Releases"

# Column indices in the current "New Releases" layout
RS_COL_TITLE = 2
RS_COL_RTYPE = 4
RS_COL_LAUNCH = 5


# ---------------------------------------------------------------------------
# Date parsing
# ---------------------------------------------------------------------------

def parse_dc_date(v):
    """DC dates are MM/DD/YYYY text strings (or real date/datetime objects)."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str) and v.strip():
        try:
            return datetime.strptime(v.strip(), '%m/%d/%Y').date()
        except ValueError:
            return None
    return None


def parse_rs_date(v):
    """
    RS Launch Date. Tolerant of:
      - real datetime/date objects (from xlsx)
      - DD/MM/YYYY strings (from a CSV export)
      - ISO YYYY-MM-DD strings
    Returns None for blanks or non-dates like 'tbc'.
    """
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Title normalisation
# ---------------------------------------------------------------------------

def normalize(s):
    s = str(s).lower().strip()
    s = re.sub(r'[^a-z0-9 ]', '', s)
    s = re.sub(r'\b(the|a|an)\b', '', s)
    return re.sub(r'\s+', ' ', s).strip()


def normalize_loose(s):
    """Also strips year suffixes like (2026) and annotations like (Ani)."""
    s = str(s).lower().strip()
    s = re.sub(r'\(\d{4}\)', '', s)
    s = re.sub(r'\([a-z]+\)', '', s)   # e.g. (Ani)
    s = re.sub(r'[^a-z0-9 ]', '', s)
    s = re.sub(r'\b(the|a|an)\b', '', s)
    return re.sub(r'\s+', ' ', s).strip()


# ---------------------------------------------------------------------------
# Load RS
# ---------------------------------------------------------------------------

def _add_rs_entry(rs, title, rtype, launch, row_no):
    """Store one RS row. The single Launch Date is used for both EST and VOD."""
    if not title or rtype not in ('Premium', 'Standard'):
        return
    entry = {'row': row_no, 'est': launch, 'vod': launch, 'title': title}
    for key in set([normalize(title), normalize_loose(title)]):
        rs[key][rtype].append(entry)


def load_rs_xlsx(xlsx_path):
    """
    Read the 'New Releases' sheet directly. Returns:
        dict: norm_title -> {release_type -> list of {row, est, vod, title}}
    Only Premium and Standard rows are kept.
    """
    rs = defaultdict(lambda: defaultdict(list))
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[RS_SHEET_NAME] if RS_SHEET_NAME in wb.sheetnames else wb.active
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) <= RS_COL_LAUNCH:
            continue
        title = str(row[RS_COL_TITLE]).strip() if row[RS_COL_TITLE] else ''
        rtype = str(row[RS_COL_RTYPE]).strip() if row[RS_COL_RTYPE] else ''
        launch = parse_rs_date(row[RS_COL_LAUNCH])
        _add_rs_entry(rs, title, rtype, launch, i)
    return rs


def load_rs_csv(csv_path):
    """
    Backward-compatible CSV reader for the current sheet layout
    (Launch Date at column index 5, used as both EST and VOD).
    """
    rs = defaultdict(lambda: defaultdict(list))
    with open(csv_path, newline='', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader, start=1):
            if len(row) <= RS_COL_LAUNCH:
                continue
            title = row[RS_COL_TITLE].strip()
            rtype = row[RS_COL_RTYPE].strip()
            launch = parse_rs_date(row[RS_COL_LAUNCH])
            _add_rs_entry(rs, title, rtype, launch, i)
    return rs


def load_rs(path):
    return load_rs_csv(path) if path.lower().endswith('.csv') else load_rs_xlsx(path)


# ---------------------------------------------------------------------------
# Load DC from XLSX
# ---------------------------------------------------------------------------

def load_dc(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = row[0]
        if not title or not str(title).strip():
            continue
        pest = parse_dc_date(row[2])
        pvod = parse_dc_date(row[4])
        est  = parse_dc_date(row[6])
        vod  = parse_dc_date(row[8])
        if not any([pest, pvod, est, vod]):
            continue
        entries.append({
            'title': str(title).strip(),
            'pest': pest, 'pvod': pvod,
            'est': est,  'vod': vod,
        })
    return entries


# ---------------------------------------------------------------------------
# Reconcile
# ---------------------------------------------------------------------------

def reconcile(dc_entries, rs):
    issues = []
    passed = []

    for dc in dc_entries:
        title = dc['title']
        nk = normalize(title)
        nk_loose = normalize_loose(title)

        rs_data = rs.get(nk) or rs.get(nk_loose)

        if rs_data is None:
            issues.append(f"MISSING FROM RS — {title}")
            continue

        entry_ok = True

        # 1. PEST / PVOD internal consistency
        if dc['pest'] and dc['pvod'] and dc['pest'] != dc['pvod']:
            issues.append(
                f"PEST/PVOD MISMATCH IN DC — {title} | "
                f"PEST={dc['pest']}, PVOD={dc['pvod']}"
            )
            entry_ok = False

        # 2. Premium check (PEST/PVOD vs RS Premium Launch Date)
        if dc['pest'] or dc['pvod']:
            premium_rows = rs_data.get('Premium', [])
            if not premium_rows:
                issues.append(f"NO PREMIUM ROW IN RS — {title}")
                entry_ok = False
            else:
                pr = premium_rows[0]
                if dc['pest'] and pr['est'] and dc['pest'] != pr['est']:
                    issues.append(
                        f"PEST ≠ RS Premium Launch — {title} | "
                        f"DC={dc['pest']}, RS={pr['est']} (row {pr['row']})"
                    )
                    entry_ok = False
                if dc['pvod'] and pr['vod'] and dc['pvod'] != pr['vod']:
                    issues.append(
                        f"PVOD ≠ RS Premium Launch — {title} | "
                        f"DC={dc['pvod']}, RS={pr['vod']} (row {pr['row']})"
                    )
                    entry_ok = False
                if dc['pvod'] and not pr['vod']:
                    issues.append(
                        f"PVOD ≠ RS Premium Launch — {title} | "
                        f"DC={dc['pvod']}, RS Launch=blank/tbc (row {pr['row']})"
                    )
                    entry_ok = False

        # 3. Standard check (EST/VOD vs RS Standard Launch Date)
        if dc['est'] or dc['vod']:
            standard_rows = rs_data.get('Standard', [])
            if not standard_rows:
                issues.append(f"NO STANDARD ROW IN RS — {title}")
                entry_ok = False
            else:
                # If multiple Standard rows, prefer the one whose date matches DC EST
                if len(standard_rows) > 1:
                    exact = [r for r in standard_rows if r['est'] == dc['est']]
                    standard_rows = exact if exact else [
                        max(standard_rows, key=lambda r: r['row'])
                    ]
                sr = standard_rows[0]
                if dc['est'] and sr['est'] and dc['est'] != sr['est']:
                    issues.append(
                        f"EST ≠ RS Standard Launch — {title} | "
                        f"DC={dc['est']}, RS={sr['est']} (row {sr['row']})"
                    )
                    entry_ok = False
                if dc['vod'] and sr['vod'] and dc['vod'] != sr['vod']:
                    issues.append(
                        f"VOD ≠ RS Standard Launch — {title} | "
                        f"DC={dc['vod']}, RS={sr['vod']} (row {sr['row']})"
                    )
                    entry_ok = False
                if dc['vod'] and not sr['vod']:
                    issues.append(
                        f"VOD ≠ RS Standard Launch — {title} | "
                        f"DC VOD={dc['vod']}, RS Launch=blank/tbc (row {sr['row']})"
                    )
                    entry_ok = False

        if entry_ok:
            passed.append(title)

    return passed, issues


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) != 3:
        print("Usage: python3 wbd_reconcile.py <digital_calendar.xlsx> <release_schedule.xlsx|.csv>")
        sys.exit(1)

    dc_path = sys.argv[1]
    rs_path = sys.argv[2]

    print(f"Loading DC:  {dc_path}")
    print(f"Loading RS:  {rs_path}")

    dc_entries = load_dc(dc_path)
    rs = load_rs(rs_path)

    total_rs = sum(len(v['Premium']) + len(v['Standard']) for v in rs.values()) // 2
    print(f"\nDC entries: {len(dc_entries)}")
    print(f"RS Premium+Standard rows loaded: {total_rs}")

    passed, issues = reconcile(dc_entries, rs)

    print(f"\n{'='*60}")
    print(f"PASS ({len(passed)})")
    print(f"{'='*60}")
    for t in passed:
        print(f"  ✅  {t}")

    print(f"\n{'='*60}")
    print(f"ISSUES ({len(issues)})")
    print(f"{'='*60}")
    if issues:
        for i, iss in enumerate(issues, 1):
            print(f"  {i}. ❌  {iss}")
    else:
        print("  ✅  No issues found!")


if __name__ == '__main__':
    main()
