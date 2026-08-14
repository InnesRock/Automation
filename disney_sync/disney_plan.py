"""
Build the full Disney calendar plan:
- All future release dates (no 28-day cap)
- EMEA comments parsed and expanded: each sub-date becomes its own day entry with
  only the countries going live that day
- Ambiguous MM/DD vs DD/MM dates flagged
"""

import warnings, re, json, zipfile, os
from datetime import datetime, date, timedelta
from collections import defaultdict

warnings.simplefilter("ignore")
import openpyxl


def load_active_threaded_comments(xlsx_path,
                                  threaded_xml="xl/threadedComments/threadedComment1.xml"):
    """Return {cell_ref: active_text} for the Film NRs sheet.

    Threaded comments form a tree: a root comment carries `done="0"`
    (active) or `done="1"` (resolved); replies have a `parentId` and no
    `done` attribute. We keep every comment whose root is active, then
    concatenate them per cell ref in document order. This bypasses
    openpyxl's `cell.comment.text`, which can surface a resolved root
    while hiding live entries — and which only returns one comment per
    cell even when multiple threads exist.
    """
    pat = re.compile(
        r'<x18tc:threadedComment\s+([^>]*?)>(.*?)</x18tc:threadedComment>',
        re.DOTALL,
    )
    text_pat = re.compile(r'<x18tc:text[^>]*>(.*?)</x18tc:text>', re.DOTALL)

    try:
        with zipfile.ZipFile(xlsx_path) as z:
            data = z.read(threaded_xml).decode("utf-8")
    except KeyError:
        return {}

    # Pass 1: collect every comment with its attrs and text (in document order).
    entries = []
    for m in pat.finditer(data):
        attrs = dict(re.findall(r'(\w+)="([^"]*)"', m.group(1)))
        tm = text_pat.search(m.group(2))
        if not tm:
            continue
        txt = (tm.group(1)
               .replace("&amp;", "&")
               .replace("&lt;", "<")
               .replace("&gt;", ">"))
        entries.append({
            "ref": attrs.get("ref", ""),
            "id": attrs.get("id"),
            "parentId": attrs.get("parentId"),
            "done": attrs.get("done"),  # only set on roots
            "text": txt,
        })

    # Pass 2: identify active root ids (done == "0").
    active_roots = {e["id"] for e in entries
                    if e.get("parentId") is None and e.get("done") == "0"}

    # Pass 3: keep each comment whose root is active; group by cell ref.
    by_ref = defaultdict(list)
    for e in entries:
        root_id = e["parentId"] or e["id"]
        if root_id not in active_roots:
            continue
        by_ref[e["ref"]].append(e["text"])

    return {ref: "\n".join(items) for ref, items in by_ref.items()}

SHEET_PATH = os.environ.get(
    "DISNEY_SHEET_PATH",
    "/sessions/cool-gallant-tesla/mnt/uploads/Disney Studio Title List & Release Schedule (3).xlsx",
)
TODAY = date.today() if os.environ.get("DISNEY_USE_TODAY") else date(2026, 6, 2)

# Category / region column map for 'Film NRs' sheet
SPECS = [
    ("Pre-Order", "NA", 5),  ("Pre-Order", "EMEA", 6),  ("Pre-Order", "FR", 7),  ("Pre-Order", "APAC", 8),
    ("Premium",  "NA", 9),   ("Premium",  "EMEA", 10),                           ("Premium",  "APAC", 11),
    ("EST",      "NA", 13),  ("EST",      "EMEA", 14),  ("EST",      "FR", 15),  ("EST",      "APAC", 16),
    ("VOD",      "NA", 18),  ("VOD",      "EMEA", 19),  ("VOD",      "FR", 20),
]

MONTHS = {m: i for i, m in enumerate(
    ["January","February","March","April","May","June",
     "July","August","September","October","November","December"], start=1)}
MONTHS.update({
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "Jun": 6, "Jul": 7,
    "Aug": 8, "Sep": 9, "Sept": 9, "Oct": 10, "Nov": 11, "Dec": 12,
})

CATEGORY_ORDER = {"Pre-Order": 1, "Premium": 2, "EST": 3, "VOD": 4}
# Display precedence for regions inside a "Regions:" line.
REGION_ORDER = {"APAC": 1, "EMEA": 2, "FR": 3, "NA": 4}

# Full EMEA region set (per user spec) used to resolve "all other regions"
# phrases in comments. Order matters for display.
EMEA_FULL = ["AT", "BE", "DK", "FI", "DE", "NL", "NO", "PL", "ES", "SE",
             "CH", "GB", "IE", "IT", "PT", "UA"]

# Tokens that look like country codes but are actually launch types or region
# names inside a comment — must never be included as "countries".
NON_COUNTRY_TOKENS = {
    "EST", "VOD", "PVOD", "POEST", "NR", "EMEA", "APAC", "NA",
    "UK",  # keep as country alias but note: UK is legit country code — exclude later
}
# Actually UK is a legit country code; restore:
NON_COUNTRY_TOKENS = {"EST", "VOD", "PVOD", "POEST", "NR", "EMEA", "APAC", "NA"}


def strip_threaded_boilerplate(text):
    """Strip the Excel threaded-comment author preamble down to content."""
    m = re.search(r"Comment:\s*(.*)", text, flags=re.DOTALL)
    if m:
        text = m.group(1)
    # collapse whitespace, convert "Reply:" boundaries to a delimiter we can split on
    text = re.sub(r"\s+", " ", text).strip()
    text = text.replace("Reply: ", " ; ").replace("Reply:", " ; ")
    return text


def parse_comment(raw, title, coord, main_cell_date, reference_year=None):
    if reference_year is None:
        reference_year = main_cell_date.year if main_cell_date else 2026
    """
    Parse an EMEA comment into [(resolved_date, countries_text, ambiguity_note)].

    Rules:
    - Numeric a/b where max>12  -> unambiguous (the >12 one is the day)
    - Numeric a/b where both<=12 -> try to infer format from any unambiguous
      date elsewhere in the same comment; else flag as ambiguous.
    - Written forms "May 5th" / "June 2nd" -> unambiguous.
    - A date token consumes the following countries-text chunk
      (letters/slashes, until next date or delimiter).
    - Country phrases like "all other regions" are preserved verbatim.
    """
    text = strip_threaded_boilerplate(raw)

    # --- First pass: find every date-like token and figure out sheet-level format
    num_dates = re.findall(r"(?<![A-Za-z0-9])(\d{1,2})/(\d{1,2})(?![A-Za-z0-9])", text)
    fmt_votes = {"MD": 0, "DM": 0}
    for a, b in num_dates:
        a, b = int(a), int(b)
        if a > 12 and b <= 12:
            fmt_votes["DM"] += 1
        elif b > 12 and a <= 12:
            fmt_votes["MD"] += 1
    # Cross-comment default: sheet is dominantly MM/DD
    sheet_default = "MD"
    local_fmt = None
    if fmt_votes["MD"] and not fmt_votes["DM"]:
        local_fmt = "MD"
    elif fmt_votes["DM"] and not fmt_votes["MD"]:
        local_fmt = "DM"
    elif fmt_votes["MD"] and fmt_votes["DM"]:
        local_fmt = None  # mixed — handle per token

    # --- Second pass: walk the tokens
    # Tokenise on whitespace after inserting spacing around key symbols
    tx = re.sub(r"[;,]", " ; ", text)
    tx = re.sub(r"–|—|-", " - ", tx)
    tx = re.sub(r"\(|\)", " ", tx)
    tokens = [t for t in tx.split() if t.strip()]

    entries = []  # list of dicts
    i = 0
    current = None
    pending_phrase = None  # phrase attaching to the NEXT date entry

    def flush():
        nonlocal current
        if current is not None and (current["countries"] or current.get("phrase")):
            entries.append(current)
        current = None

    while i < len(tokens):
        tok = tokens[i]

        # Numeric date
        m = re.fullmatch(r"(\d{1,2})/(\d{1,2})", tok)
        if m:
            flush()
            a, b = int(m.group(1)), int(m.group(2))
            ambiguity = None
            if a > 12 and b <= 12:
                mo, da, fmt = b, a, "DD/MM"
            elif b > 12 and a <= 12:
                mo, da, fmt = a, b, "MM/DD"
            else:
                # both <=12
                if local_fmt == "MD":
                    mo, da, fmt = a, b, "MM/DD (inferred from comment)"
                elif local_fmt == "DM":
                    mo, da, fmt = b, a, "DD/MM (inferred from comment)"
                else:
                    # fallback to sheet default, but prefer the future interpretation
                    # if one is past and the other is future
                    md_date = dm_date = None
                    try: md_date = date(reference_year, a, b)
                    except ValueError: pass
                    try: dm_date = date(reference_year, b, a)
                    except ValueError: pass
                    md_future = md_date and md_date >= TODAY
                    dm_future = dm_date and dm_date >= TODAY
                    if md_future and not dm_future:
                        mo, da, fmt = a, b, "MM/DD (only future interpretation)"
                    elif dm_future and not md_future:
                        mo, da, fmt = b, a, "DD/MM (only future interpretation)"
                    else:
                        mo, da = (a, b) if sheet_default == "MD" else (b, a)
                        fmt = "ambiguous (both interpretations valid)"
                        ambiguity = (f"'{tok}' is ambiguous: parsed as "
                                     f"{mo:02d}-{da:02d} (MM/DD); could also be "
                                     f"{b:02d}-{a:02d} (DD/MM)")
            try:
                d = date(reference_year, mo, da)
            except ValueError:
                d = None
            current = {"raw_date": tok, "date": d, "fmt": fmt, "countries": [],
                       "ambiguity": ambiguity, "phrase": pending_phrase}
            pending_phrase = None
            i += 1
            continue

        # Written date: Month + Day
        if tok in MONTHS and i + 1 < len(tokens):
            day_tok = re.sub(r"(st|nd|rd|th)$", "", tokens[i+1])
            if day_tok.isdigit():
                flush()
                d = date(reference_year, MONTHS[tok], int(day_tok))
                current = {"raw_date": f"{tok} {tokens[i+1]}", "date": d,
                           "fmt": "written", "countries": [], "ambiguity": None,
                           "phrase": pending_phrase}
                pending_phrase = None
                i += 2
                continue

        # Written date: Day + Month (e.g. "18 Aug", "8th September")
        day_tok = re.sub(r"(st|nd|rd|th)$", "", tok)
        if day_tok.isdigit() and i + 1 < len(tokens) and tokens[i+1] in MONTHS:
            flush()
            d = date(reference_year, MONTHS[tokens[i+1]], int(day_tok))
            current = {"raw_date": f"{tok} {tokens[i+1]}", "date": d,
                       "fmt": "written", "countries": [], "ambiguity": None,
                       "phrase": pending_phrase}
            pending_phrase = None
            i += 2
            continue

        # "all other regions" phrase.
        # Behaviour:
        # - If the current entry already has a date and countries (e.g.
        #   "May 5th IE/IT/PT/UA all other regions June 2nd"), the phrase
        #   applies to the NEXT date entry — store as pending_phrase.
        # - If the current entry has a date but no countries yet
        #   (e.g. "June 2nd all other regions"), attach to current.
        # - If no current entry, also store as pending for the next date.
        if tok.lower() == "all" and i + 2 < len(tokens) and tokens[i+1].lower() == "other" and tokens[i+2].lower() == "regions":
            if current is not None and not current["countries"] and not current.get("phrase"):
                current["phrase"] = "all other regions"
            else:
                pending_phrase = "all other regions"
            i += 3
            continue

        # "Everywhere except X/Y [country codes...]" -- drop the "Everywhere
        # except X/Y" prefix and continue parsing the trailing country codes
        # normally. The phrase itself is discarded (user request: don't show
        # "Everywhere except Italy/France" in the description).
        if tok.lower() == "everywhere":
            if current is None:
                current = {"raw_date": None, "date": main_cell_date, "fmt": "main cell",
                           "countries": [], "ambiguity": None, "phrase": None}
            i += 1  # past "everywhere"
            if i < len(tokens) and tokens[i].lower() == "except":
                i += 1  # past "except"
            # Consume the exception region token (e.g. "Italy/France")
            if i < len(tokens):
                i += 1
            continue

        # Country-code chunk (slash-separated uppercase codes, possibly with "/")
        if re.fullmatch(r"[A-Z]{2,4}(?:/[A-Z]{2,4})*", tok):
            if current is None:
                # no preceding date -> attach to main cell date
                current = {"raw_date": None, "date": main_cell_date, "fmt": "main cell",
                           "countries": [], "ambiguity": None, "phrase": None}
            # filter out type/region tokens that aren't actual country codes
            parts = [p for p in tok.split("/") if p not in NON_COUNTRY_TOKENS]
            # flag a known typo: "PLES" (missing slash between PL and ES)
            expanded = []
            for p in parts:
                if p == "PLES":
                    expanded.extend(["PL", "ES"])
                    current.setdefault("_notes", []).append(
                        f"source typo 'PLES' expanded to 'PL/ES'")
                else:
                    expanded.append(p)
            current["countries"].extend(expanded)
            i += 1
            continue

        # delimiters / noise
        i += 1

    flush()
    return entries


def main():
    wb = openpyxl.load_workbook(SHEET_PATH, data_only=True)
    ws = wb["Film NRs"]

    # Pass 1: collect all future release entries from main cells
    # plus catalogue EMEA comments for later expansion
    main_entries = []      # (date, cat, region, title, media)
    emea_cells = {}        # (title, cat, row) -> {"date": main_cell_date, "comment": raw_text, "coord": ...}
    row_year = {}          # row -> inferred year context from row's datetime cells

    for r in range(3, ws.max_row + 1):
        # Skip hidden rows — these are historical entries and can always be ignored
        if ws.row_dimensions[r].hidden:
            continue
        title = ws.cell(row=r, column=1).value
        media = ws.cell(row=r, column=4).value
        if not title or str(title).strip() == "":
            continue

        # Determine row's year context from any datetime cell in the row
        years_seen = []
        for _, _, col in SPECS:
            v = ws.cell(row=r, column=col).value
            if isinstance(v, datetime):
                years_seen.append(v.year)
        if years_seen:
            row_year[r] = max(years_seen)  # use latest year seen in the row

        for cat, region, col in SPECS:
            cell = ws.cell(row=r, column=col)
            v = cell.value
            if isinstance(v, datetime):
                d = v.date()
                if d >= TODAY:
                    main_entries.append((d, cat, region, title, media, r, col))
                # also register EMEA cells (even if past) so the comment expansion can
                # pick up future sub-dates that live inside a past main cell's comment
            if region == "EMEA" and cell.comment is not None:
                emea_cells[(title, cat, r, col)] = {
                    "main_date": v.date() if isinstance(v, datetime) else None,
                    "raw_comment": cell.comment.text,
                    "coord": cell.coordinate,
                    "row_year": row_year.get(r),
                }

    # Pass 2: parse EMEA comments into sub-entries (date + specific countries)
    # Each EMEA cell with a comment yields a *list* of (date, countries_text) pairs.
    # We will replace the simple (date, "EMEA") main-cell entry with expanded
    # (date, "EMEA (countries)") entries.
    comment_expansions = {}  # (title, cat, row) -> list of parsed entries
    ambiguities = []

    for key, info in emea_cells.items():
        title, cat, row, col = key
        # Prefer the main cell's year, else the row's inferred year, else 2026
        ref_year = None
        if info["main_date"]:
            ref_year = info["main_date"].year
        elif info["row_year"]:
            ref_year = info["row_year"]
        parsed = parse_comment(info["raw_comment"], title, info["coord"],
                               info["main_date"], reference_year=ref_year)
        comment_expansions[key] = parsed
        for e in parsed:
            if e["ambiguity"]:
                ambiguities.append({
                    "title": title, "cat": cat, "coord": info["coord"],
                    "raw_date": e["raw_date"],
                    "resolved": e["date"].isoformat() if e["date"] else None,
                    "note": e["ambiguity"],
                    "countries": "/".join(e["countries"]) or e["phrase"] or "",
                })
            # also flag DD/MM when sheet is dominantly MM/DD
            if e["fmt"] == "DD/MM":
                ambiguities.append({
                    "title": title, "cat": cat, "coord": info["coord"],
                    "raw_date": e["raw_date"],
                    "resolved": e["date"].isoformat() if e["date"] else None,
                    "note": f"{e['raw_date']} uses DD/MM format (rest of sheet uses MM/DD)",
                    "countries": "/".join(e["countries"]) or e["phrase"] or "",
                })

    # Pass 3: build the expanded release list
    # For each EMEA main entry, replace with the comment sub-entries (if any comment exists).
    # For non-EMEA entries, pass through unchanged.
    releases = []  # dicts with date, title, cat, region_label, country_detail, source

    # Build a quick lookup to know which (title, cat, row) have comment expansions
    emea_key_by_row = {}
    for key in comment_expansions:
        title, cat, row, col = key
        emea_key_by_row[(title, cat, row)] = key

    for (d, cat, region, title, media, r, col) in main_entries:
        if region == "EMEA" and (title, cat, r) in emea_key_by_row:
            # Skip — will be handled by comment expansion below
            continue
        releases.append({
            "date": d, "title": title, "cat": cat,
            "region_label": region,
            "country_detail": None,
            "media": media,
        })

    # --- Per-cell overrides (user-directed corrections to source data) ---
    # Send Help VOD S150: the "5/12 BE/DK/FI/IT/NL/NO/SE" sub-date in the comment
    # is outdated — user confirmed all EMEA countries should launch on 5/19 with
    # the main cell date. Collapse comment sub-dates back onto the main date.
    COMMENT_OVERRIDES_TO_MAIN_DATE = {
        ("Send Help", "VOD", "S150"),
    }

    # Add expanded EMEA entries
    for key, parsed in comment_expansions.items():
        title, cat, row, col = key
        coord = ws.cell(row=row, column=col).coordinate
        # Use title_media if available (defined in emit_json); otherwise fall
        # back to direct lookup (main()).
        try:
            media = title_media.get(title) or ws.cell(row=row, column=4).value
        except NameError:
            media = ws.cell(row=row, column=4).value
        main_date = emea_cells[key]["main_date"]

        # If this (title, cat, coord) is overridden, replace all parsed sub-dates
        # with a single entry on the main cell date using the union of countries.
        if (title, cat, coord) in COMMENT_OVERRIDES_TO_MAIN_DATE and main_date:
            union_countries = []
            for e in parsed:
                union_countries.extend(e["countries"])
            detail = "/".join(union_countries) if union_countries else None
            if main_date >= TODAY:
                releases.append({
                    "date": main_date, "title": title, "cat": cat,
                    "region_label": "EMEA",
                    "country_detail": detail,
                    "media": media,
                })
            continue

        for e in parsed:
            if not e["date"] or e["date"] < TODAY:
                continue
            countries = "/".join(e["countries"]) if e["countries"] else None
            if e.get("phrase"):
                # keep the phrase as the detail
                phrase = e["phrase"]
                detail = f"{countries} ({phrase})" if countries else phrase
            else:
                detail = countries
            releases.append({
                "date": e["date"], "title": title, "cat": cat,
                "region_label": "EMEA",
                "country_detail": detail,
                "media": media,
            })

    # Pass 4a: merge entries that share (date, title, cat, country_detail, media)
    # by combining their region labels. E.g. Ready Or Not 2 Pre-Order NA + APAC on
    # 30 Apr should be a single bullet "Regions: APAC, NA".
    merged = defaultdict(list)
    for rel in releases:
        key = (rel["date"], rel["title"], rel["cat"],
               rel["country_detail"], rel.get("media"))
        merged[key].append(rel["region_label"])

    # Pass 4b: dedupe same-day (title, cat, media) pairs where one row is FR and
    # another row is EMEA with country_detail == "FR". Collapse to a single FR entry.
    dedup = {}  # (date, title, cat, media) -> list of rows
    prelim = []
    for (d, title, cat, detail, media), regions in merged.items():
        prelim.append({
            "date": d, "title": title, "cat": cat,
            "region_label": ", ".join(sorted(set(regions))),
            "country_detail": detail, "media": media,
        })

    by_day = defaultdict(list)
    # Build a lookup of FR rows per (date, title, cat, media) so we can suppress
    # the EMEA row when it only restates FR.
    fr_keys = {
        (p["date"], p["title"], p["cat"], p.get("media"))
        for p in prelim if "FR" in p["region_label"].split(", ")
    }
    for p in prelim:
        regions = p["region_label"].split(", ")
        key = (p["date"], p["title"], p["cat"], p.get("media"))
        # Suppress an EMEA-only row whose country_detail is exactly "FR"
        # when there's already a standalone FR row on the same day/title/cat.
        if (regions == ["EMEA"]
                and p["country_detail"] == "FR"
                and key in fr_keys):
            continue
        by_day[p["date"]].append(p)

    # Sort each day
    for d in by_day:
        by_day[d].sort(key=lambda x: (CATEGORY_ORDER.get(x["cat"], 99), x["title"]))

    # Print plan
    all_days = sorted(by_day.keys())
    print("=" * 70)
    print(f"DISNEY CALENDAR PLAN — generated {TODAY}")
    print("=" * 70)
    print(f"\nWindow: {TODAY} -> {all_days[-1] if all_days else 'n/a'}  (no cap)")
    print(f"Distinct days with releases: {len(all_days)}")
    print(f"Total release entries (post comment-expansion): {sum(len(v) for v in by_day.values())}")
    print()

    for d in all_days:
        day_releases = by_day[d]
        print(f"--- {d.strftime('%a %Y-%m-%d')}  ({len(day_releases)} release(s)) ---")
        for r in day_releases:
            line = f"  [{r['cat']}] {r['title']} — {r['region_label']}"
            if r["country_detail"]:
                line += f"   EMEA detail: {r['country_detail']}"
            if r["media"] and r["media"] != "Film":
                line += f"   (Media: {r['media']})"
            print(line)
        print()

    # Ambiguities — filter to only those affecting in-window (future) dates,
    # and drop any coord that was resolved via COMMENT_OVERRIDES_TO_MAIN_DATE.
    override_coords = {coord for (_t, _c, coord) in COMMENT_OVERRIDES_TO_MAIN_DATE}
    in_window = [a for a in ambiguities
                 if a["resolved"]
                 and date.fromisoformat(a["resolved"]) >= TODAY
                 and a["coord"] not in override_coords]

    print("=" * 70)
    print(f"AMBIGUITY / FORMAT FLAGS (in-window only: {len(in_window)})")
    print("=" * 70)
    if not in_window:
        print("None.")
    for a in in_window:
        print(f"  • {a['title']} / {a['cat']} @ {a['coord']} — '{a['raw_date']}'")
        print(f"    Resolved to: {a['resolved']}  (countries: {a['countries']})")
        print(f"    Note: {a['note']}")
        print()

    past_count = len(ambiguities) - len(in_window)
    if past_count:
        print(f"({past_count} additional flags are on historical/past-date rows "
              "and don't affect the plan; omitted.)")


def load_tv_releases(xlsx_path):
    """Load future TV releases from the 'TV NRs' sheet.

    Schema (row 3 headers):
        col 1: Vendor Identifier
        col 2: Upcoming Releases (title)
        col 3: Film/TV  (= "TV")
        col 5: Day      (release datetime)
        col 7: Season Number
        col 8: Regions  (free-form string, e.g. "US", "DE", "US, CA")
    Hidden rows are skipped (historical).

    Returns a list of dicts: {date, title, season, regions}.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "TV NRs" not in wb.sheetnames:
        return []
    ws = wb["TV NRs"]
    out = []
    for r in range(4, ws.max_row + 1):
        if ws.row_dimensions[r].hidden:
            continue
        title = ws.cell(row=r, column=2).value
        day = ws.cell(row=r, column=5).value
        season = ws.cell(row=r, column=7).value
        regions = ws.cell(row=r, column=8).value
        if not title or not isinstance(day, datetime):
            continue
        d = day.date()
        if d < TODAY:
            continue
        out.append({
            "date": d,
            "title": str(title).strip(),
            "season": int(season) if isinstance(season, (int, float)) else None,
            "regions": str(regions).strip() if regions else None,
        })
    return out


def route_tv_date(tv_date, film_anchor_dates):
    """Return the calendar-invite date a TV release should land on.

    Day-of-week rules:
      - Tue, Wed, Fri        → keep on natural day
      - Mon                  → keep if a Film/Bundle is on the same Mon; else push to Tue
      - Thu                  → keep if a Film/Bundle is on the same Thu; else push to Fri
      - Sat, Sun             → target is the *following* Tue, unless the intervening
                               Mon has a Film/Bundle anchor (then merge into Mon)
    """
    weekday = tv_date.weekday()  # 0=Mon..6=Sun

    if weekday == 0:  # Mon
        if tv_date in film_anchor_dates:
            return tv_date
        return tv_date + timedelta(days=1)  # Tue

    if weekday == 3:  # Thu
        if tv_date in film_anchor_dates:
            return tv_date
        return tv_date + timedelta(days=1)  # Fri

    if weekday in (5, 6):  # Sat, Sun
        days_to_tue = (1 - weekday) % 7  # Sat→3, Sun→2
        next_tue = tv_date + timedelta(days=days_to_tue)
        next_mon = next_tue - timedelta(days=1)
        if next_mon in film_anchor_dates:
            return next_mon
        return next_tue

    return tv_date  # Tue/Wed/Fri


def emit_json():
    """Emit the plan as JSON so the MCP write layer can consume it."""
    import json as _json
    # Re-run main() logic but return the structured data
    wb = openpyxl.load_workbook(SHEET_PATH, data_only=True)
    ws = wb["Film NRs"]

    # Pull active (un-resolved) threaded comments from the raw xlsx XML.
    # Resolved threads are deliberately ignored — they're stale notes.
    active_comments = load_active_threaded_comments(SHEET_PATH)

    # Build merge map so slave cells inherit the master value/comment. Some
    # rows (e.g. bundle rows) share cells with the film row above via vertical
    # merges, and openpyxl only exposes the value on the master cell.
    merge_master = {}
    for mr in ws.merged_cells.ranges:
        for rr in range(mr.min_row, mr.max_row + 1):
            for cc in range(mr.min_col, mr.max_col + 1):
                merge_master[(rr, cc)] = (mr.min_row, mr.min_col)

    def cell_value(rr, cc):
        v = ws.cell(row=rr, column=cc).value
        if v is None and (rr, cc) in merge_master:
            mr_r, mr_c = merge_master[(rr, cc)]
            v = ws.cell(row=mr_r, column=mr_c).value
        return v

    def cell_comment(rr, cc):
        c = ws.cell(row=rr, column=cc).comment
        if c is None and (rr, cc) in merge_master:
            mr_r, mr_c = merge_master[(rr, cc)]
            c = ws.cell(row=mr_r, column=mr_c).comment
        return c

    main_entries = []
    emea_cells = {}
    row_year = {}

    title_type_flag = {}  # title -> value of column C (Type: "4K", "Cat", "NR")
    title_media = {}      # title -> value of column D (e.g. "Film", "Bundle")
    title_row_order = {}  # title -> first row it appears on (source sheet order)

    for r in range(3, ws.max_row + 1):
        if ws.row_dimensions[r].hidden:
            continue
        title = ws.cell(row=r, column=1).value
        media = ws.cell(row=r, column=4).value
        type_flag = ws.cell(row=r, column=3).value
        if not title or str(title).strip() == "":
            continue
        if type_flag:
            title_type_flag[title] = str(type_flag).strip()
        if media:
            title_media[title] = media
        title_row_order.setdefault(title, r)
        years_seen = []
        for _, _, col in SPECS:
            v = cell_value(r, col)
            if isinstance(v, datetime):
                years_seen.append(v.year)
        if years_seen:
            row_year[r] = max(years_seen)

        for cat, region, col in SPECS:
            v = cell_value(r, col)
            coord = ws.cell(row=r, column=col).coordinate
            # Prefer ACTIVE threaded-comment text over openpyxl's
            # cell.comment.text (which can surface resolved threads).
            # Honor vertical merges: a slave cell (e.g. bundle row 159)
            # should inherit the master cell's comment text (row 158).
            active_text = active_comments.get(coord)
            if active_text is None and (r, col) in merge_master:
                mr_r, mr_c = merge_master[(r, col)]
                master_coord = ws.cell(row=mr_r, column=mr_c).coordinate
                active_text = active_comments.get(master_coord)
            if isinstance(v, datetime):
                d = v.date()
                if d >= TODAY:
                    main_entries.append((d, cat, region, title, media, r, col))
            if region == "EMEA" and active_text:
                emea_cells[(title, cat, r, col)] = {
                    "main_date": v.date() if isinstance(v, datetime) else None,
                    "raw_comment": active_text,
                    "coord": coord,
                    "row_year": row_year.get(r),
                }

    comment_expansions = {}
    for key, info in emea_cells.items():
        title, cat, row, col = key
        ref_year = None
        if info["main_date"]:
            ref_year = info["main_date"].year
        elif info["row_year"]:
            ref_year = info["row_year"]
        parsed = parse_comment(info["raw_comment"], title, info["coord"],
                               info["main_date"], reference_year=ref_year)

        # Resolve "all other regions" phrase: compute EMEA_FULL minus the
        # union of explicit country codes from the cell's other entries.
        union_non_phrase = set()
        for e in parsed:
            if e.get("phrase") != "all other regions":
                union_non_phrase.update(e["countries"])
        for e in parsed:
            if e.get("phrase") == "all other regions":
                resolved = [c for c in EMEA_FULL if c not in union_non_phrase]
                e["countries"] = resolved
                e["phrase"] = None  # phrase has been materialised

        comment_expansions[key] = parsed

    releases = []
    emea_key_by_row = {(t, c, r): (t, c, r, col) for (t, c, r, col) in comment_expansions.keys()}

    for (d, cat, region, title, media, r, col) in main_entries:
        if region == "EMEA" and (title, cat, r) in emea_key_by_row:
            continue
        releases.append({"date": d, "title": title, "cat": cat,
                         "region_label": region, "country_detail": None, "media": media})

    COMMENT_OVERRIDES_TO_MAIN_DATE = {("Send Help", "VOD", "S150")}

    for key, parsed in comment_expansions.items():
        title, cat, row, col = key
        coord = ws.cell(row=row, column=col).coordinate
        # Use title_media if available (defined in emit_json); otherwise fall
        # back to direct lookup (main()).
        try:
            media = title_media.get(title) or ws.cell(row=row, column=4).value
        except NameError:
            media = ws.cell(row=row, column=4).value
        main_date = emea_cells[key]["main_date"]

        if (title, cat, coord) in COMMENT_OVERRIDES_TO_MAIN_DATE and main_date:
            union_countries = []
            for e in parsed:
                union_countries.extend(e["countries"])
            detail = "/".join(union_countries) if union_countries else None
            if main_date >= TODAY:
                releases.append({"date": main_date, "title": title, "cat": cat,
                                 "region_label": "EMEA", "country_detail": detail,
                                 "media": media})
            continue

        emitted = 0
        for e in parsed:
            if not e["date"] or e["date"] < TODAY:
                continue
            countries = "/".join(e["countries"]) if e["countries"] else None
            if e.get("phrase"):
                phrase = e["phrase"]
                detail = f"{countries} ({phrase})" if countries else phrase
            else:
                detail = countries
            releases.append({"date": e["date"], "title": title, "cat": cat,
                             "region_label": "EMEA", "country_detail": detail,
                             "media": media})
            emitted += 1

        # Fallback: comment had no usable dated entries (e.g. just an
        # informational note like "after NA"). Use the main-cell date with
        # no country detail so the launch isn't dropped.
        if emitted == 0 and main_date and main_date >= TODAY:
            releases.append({"date": main_date, "title": title, "cat": cat,
                             "region_label": "EMEA", "country_detail": None,
                             "media": media})

    # Merge + dedupe
    merged = defaultdict(list)
    for rel in releases:
        key = (rel["date"], rel["title"], rel["cat"],
               rel["country_detail"], rel.get("media"))
        merged[key].append(rel["region_label"])

    prelim = []
    for (d, title, cat, detail, media), regions in merged.items():
        prelim.append({"date": d, "title": title, "cat": cat,
                       "region_label": ", ".join(sorted(set(regions))),
                       "country_detail": detail, "media": media})

    fr_keys = {(p["date"], p["title"], p["cat"], p.get("media"))
               for p in prelim if "FR" in p["region_label"].split(", ")}
    by_day = defaultdict(list)
    for p in prelim:
        regions = p["region_label"].split(", ")
        key = (p["date"], p["title"], p["cat"], p.get("media"))
        if (regions == ["EMEA"] and p["country_detail"] == "FR" and key in fr_keys):
            continue
        by_day[p["date"]].append(p)

    for d in by_day:
        by_day[d].sort(key=lambda x: (CATEGORY_ORDER.get(x["cat"], 99), x["title"]))

    # --- TV NRs ingestion + routing ---------------------------------------
    # Pull TV releases from the 'TV NRs' sheet and route them by day-of-week:
    #   • Mon/Thu TV stays put iff a Film/Bundle release shares the day; else
    #     pushes to the following Tue/Fri.
    #   • Sat/Sun TV targets the next Tue, except an intervening Mon with a
    #     Film/Bundle anchor catches it.
    # Routed TV entries are merged into by_day_tv (separate from Film entries
    # so we can render them with the "(TV)" suffix and skip the Type bullet).
    film_anchor_dates = set(by_day.keys())
    tv_releases = load_tv_releases(SHEET_PATH)
    by_day_tv = defaultdict(list)
    for tv in tv_releases:
        target = route_tv_date(tv["date"], film_anchor_dates)
        by_day_tv[target].append(tv)

    # Final grouping for rendering:
    # Outer:  one <li> per title on the day
    # Inner:  one "Regions: ..." sub-bullet per unique (cat-combo, detail)
    # Cats that share the same (regions, detail) are combined (e.g. EST/VOD).
    # 4K flag prefixes the Type string.
    out = []
    all_dates = sorted(set(by_day.keys()) | set(by_day_tv.keys()))
    for d in all_dates:
        day_releases = by_day.get(d, [])
        day_tv = by_day_tv.get(d, [])

        # Step 1: per (title, cat) collapse regions + detail
        per_title_cat = {}
        for r in day_releases:
            key = (r["title"], r["cat"])
            if key not in per_title_cat:
                per_title_cat[key] = {"regions": set(), "detail": None}
            for reg in r["region_label"].split(", "):
                per_title_cat[key]["regions"].add(reg)
            if r.get("country_detail"):
                existing = per_title_cat[key]["detail"]
                if existing and existing != r["country_detail"]:
                    per_title_cat[key]["detail"] = existing + "; " + r["country_detail"]
                elif not existing:
                    per_title_cat[key]["detail"] = r["country_detail"]

        # Step 2: per title, group regions by their cat-set so a region that
        # gets multiple cats becomes part of a combined "EST/VOD" Type line,
        # even when its cat-set spans a different region set than another cat.
        # Example: Devil Wears Prada 6/9 — 4K EST is in {APAC, EMEA, FR} and
        # 4K VOD is in {EMEA, FR}. Region-by-region:
        #     APAC → {EST};  EMEA → {EST, VOD};  FR → {EST, VOD}
        # Groups by cat-set → "APAC: EST" + "EMEA, FR: EST/VOD".
        # (Previous logic grouped by region-set first, which prevented the
        # collapse since {APAC, EMEA, FR} ≠ {EMEA, FR}.)
        by_title = defaultdict(lambda: {"subgroups": {}, "order": []})

        # Build {title -> {(region, detail) -> set(cats)}}. detail only applies
        # to EMEA; non-EMEA regions always carry detail=None. Different EMEA
        # details (e.g. different country lists for EST vs VOD) stay as
        # distinct keys so they don't merge incorrectly.
        # NOTE: avoid shadowing the outer-loop `d` (the day's date). Use
        # explicit names like `detail_str` and `regions_set` inside.
        per_title_region_cats = defaultdict(lambda: defaultdict(set))
        for (title, cat), info in per_title_cat.items():
            for reg in info["regions"]:
                detail_str = info["detail"] if reg == "EMEA" else None
                per_title_region_cats[title][(reg, detail_str)].add(cat)

        for title, region_detail_cats in per_title_region_cats.items():
            # Group (region, detail) by frozenset(cats)
            groups_by_catset = defaultdict(list)
            for (reg, detail_str), cats in region_detail_cats.items():
                groups_by_catset[frozenset(cats)].append((reg, detail_str))

            for catset, region_details in groups_by_catset.items():
                regions_set = {reg for (reg, _) in region_details}
                # Only EMEA carries detail; pick it up from the EMEA entry if any.
                emea_detail = next(
                    (ds for (reg, ds) in region_details if reg == "EMEA" and ds),
                    None,
                )
                sg_key = (frozenset(regions_set), emea_detail)
                sg = by_title[title]["subgroups"].get(sg_key)
                if sg is None:
                    sg = {"regions": regions_set, "detail": emea_detail,
                          "cats": list(catset)}
                    by_title[title]["subgroups"][sg_key] = sg
                    by_title[title]["order"].append(sg_key)

        # Order titles: by earliest (smallest) cat precedence present,
        # then by the title's row order in the source spreadsheet (so related
        # titles, e.g. a film and its own bundle, stay adjacent as they do
        # in the sheet, instead of scattering alphabetically).
        title_order = sorted(
            by_title.keys(),
            key=lambda t: (
                min(
                    CATEGORY_ORDER.get(c, 99)
                    for sg in by_title[t]["subgroups"].values()
                    for c in sg["cats"]
                ),
                title_row_order.get(t, 10**9),
            ),
        )

        count = len(title_order) + len(day_tv)
        lines = [f"<p><b>Disney releases landing today ({count}):</b></p>", "<ul>"]
        for title in title_order:
            media = title_media.get(title, "Film")
            flag = title_type_flag.get(title)

            # Title header: append "(Bundle)" etc. when media != Film
            header = f"<b>{title}</b>"
            if media and media != "Film":
                header = f"<b>{title}</b> ({media})"

            # Sub-groups sorted by smallest CAT precedence inside each sub-group
            subgroups = list(by_title[title]["subgroups"].values())
            subgroups.sort(key=lambda sg: min(
                CATEGORY_ORDER.get(c, 99) for c in sg["cats"]))

            lines.append(f"<li>{header}")
            lines.append("<ul>")
            for sg in subgroups:
                regions_sorted = sorted(
                    sg["regions"],
                    key=lambda rr: REGION_ORDER.get(rr, 99),
                )
                # France lives in its own column in the source sheet, so
                # "EMEA" never includes FR by default. Annotate EMEA
                # depending on whether France is going live the same day:
                #   • FR co-listed in this sub-group → "EMEA (incl. FR)"
                #     AND drop the standalone "FR" from the line (merged
                #     into the suffix).
                #   • FR not co-listed, no EMEA country detail shown →
                #     "EMEA (excl. FR)" (makes France's absence explicit).
                #   • FR not co-listed, EMEA country detail shown → just
                #     "EMEA" (the country list itself makes scope clear).
                has_emea_detail = bool(sg["detail"])
                has_emea = "EMEA" in sg["regions"]
                fr_co_listed = has_emea and "FR" in sg["regions"]

                def _render_region(rr):
                    if rr == "EMEA":
                        if fr_co_listed:
                            return "EMEA (incl. FR)"
                        if not has_emea_detail:
                            return "EMEA (excl. FR)"
                    return rr

                regions_str = ", ".join(
                    _render_region(rr)
                    for rr in regions_sorted
                    if not (rr == "FR" and fr_co_listed)
                )
                cats_sorted = sorted(set(sg["cats"]),
                                     key=lambda c: CATEGORY_ORDER.get(c, 99))
                type_str = "/".join(cats_sorted)
                if flag == "4K":
                    type_str = f"4K {type_str}"

                lines.append(f"<li>Regions: {regions_str}")
                lines.append("<ul>")
                lines.append(f"<li>Type: {type_str}</li>")
                if sg["detail"]:
                    lines.append(f"<li>EMEA regions: {sg['detail']}</li>")
                lines.append("</ul>")
                lines.append("</li>")
            lines.append("</ul>")
            lines.append("</li>")
        # TV entries (rendered after Film/Bundle titles on this day).
        # The source title already carries the season suffix (e.g. "Foo, Season 1");
        # we just append a "(TV)" tag and skip the EST/VOD/Pre-Order Type line.
        for tv in sorted(day_tv, key=lambda t: t["title"]):
            lines.append(f"<li><b>{tv['title']}</b> (TV)")
            lines.append("<ul>")
            if tv.get("regions"):
                lines.append(f"<li>Regions: {tv['regions']}</li>")
            lines.append("</ul>")
            lines.append("</li>")
        lines.append("</ul>")
        out.append({"date": d.isoformat(), "description": "".join(lines),
                    "count": count})

    print(_json.dumps(out, indent=2))


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "emit":
        emit_json()
    else:
        main()
